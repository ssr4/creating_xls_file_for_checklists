#!/usr/bin/python3

from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import os
import check_date
import db as database
import curl_request as curl
DIR = '/home/feodor/Desktop/programming/rzd/php-server-for-loading/orders/cusi/'
FORMAT_DATE = '%Y-%m-%d %H:%M:%S'


def connect_to_db():
    return database.DB(user="mptablo",
                       # пароль, который указали при установке PostgreSQL
                       password="tablo_rkbtyn",
                       host="localhost",
                       port="5432",
                       db_name="tablo"
                       )


def adjust_font():
    return Font(name='Arial', size=12, italic=True)


def adjust_alignment(excel_cell):
    excel_cell.alignment = Alignment(
        horizontal='justify', vertical='top', indent=100.5, readingOrder=2.5,)
    excel_cell.alignment = Alignment(wrap_text=True)


def get_cell(ws, row, column):
    return ws.cell(row=row,
                   column=column)


def read_and_create_file(file_name, checklist_data, operational_order, date_of_operational_order):
    wb = load_workbook('checklist_example.xlsx')
    ws = wb.active
    # Название дирекции | ФИО | Дата, время формирования
    min_row = 5
    today = datetime.today().strftime(FORMAT_DATE)
    ws['C4'] = f'{operational_order} от {date_of_operational_order}'
    for row in ws.iter_rows(min_row=min_row, min_col=3, max_col=5, max_row=min_row, values_only=True):
        # responsible manager
        ws[f'C{min_row}'] = f"{checklist_data['service_name']}"
        # chief
        ws[f'D{min_row}'] = f"ФИО отвественного руководителя: {checklist_data['boss']}"
        # date of formation
        ws[f'E{min_row}'] = f'{today}'
    wb.save(f'./actions/{file_name}_{today}.xlsx')
    min_row = 7
    max_row = min_row + len(checklist_data['regions']) - 1
    min_col = 1
    max_col = 8
    # проходим по каждой строке
    for idx, row in enumerate(ws.iter_rows(min_row=min_row, max_row=max_row)):
        # проходим по каждому столбцу
        for idx_col in range(min_col, max_col):
            if (idx_col == 1):
                get_cell(ws, idx + min_row, idx_col).value = f'{idx + 1}'
            if (idx_col == 2):
                get_cell(ws, idx + min_row,
                         idx_col).value = checklist_data['regions'][idx]
            if (idx_col == 3):
                get_cell(ws, idx + min_row,
                         idx_col).value = checklist_data['weather_condition'][idx]
            if (idx_col == 4):
                get_cell(ws, idx + min_row,
                         idx_col).value = checklist_data['action'][idx]
            if (idx_col == 5):
                pass
            if (idx_col == 6):
                get_cell(ws, idx + min_row,
                         idx_col).value = checklist_data['status'][idx]
            if (idx_col == 7):
                get_cell(ws, idx + min_row,
                         idx_col).value = checklist_data['full_name'][idx]
            get_cell(ws, idx + min_row,
                     idx_col).font = adjust_font()
            adjust_alignment(get_cell(ws, idx + min_row,
                                      idx_col))
    wb.save(f'./actions/{file_name}_{today}.xlsx')


try:
    files = [f for f in os.listdir(DIR)]
    db = connect_to_db()
    for file in files:
        is_operational_file_received = check_date.check_date(db, file)
        if is_operational_file_received:
            upload_time = datetime.fromtimestamp(os.path.getctime(
                f'{DIR}{file}'))
            # print(upload_time)
            services = db.query_with_a_dictionary(
                'select * from accounts.get_all_services', '').fetchall()
            # проходим по всем службам
            for service in services:
                sl_code = service['id']
                # получаем данные по всем мероприятиям для одной службы в формате dictionary
                data_for_a_single_service = db.query_with_a_dictionary(
                    'select * from tablo_content.actual_storm_actions_for_excel where sl_code=%s',
                    (sl_code,)).fetchall()
                if (len(data_for_a_single_service)):
                    # наименование файла
                    file_name = f"{data_for_a_single_service[0]['sl_name']}_Чеклист"
                    # словарь для хранения данных по службам
                    recording_service_data = dict()
                    recording_service_data['service_name'] = data_for_a_single_service[0]['sl_full_name']
                    # Ответственный исполнитель
                    recording_service_data['boss'] = data_for_a_single_service[0]['fullname']
                    recording_service_data['regions'] = list()
                    recording_service_data['weather_condition'] = list()
                    recording_service_data['action'] = list()
                    recording_service_data['status'] = list()
                    recording_service_data['full_name'] = list()
                    for full_service_data in data_for_a_single_service:
                        recording_service_data['regions'].append(
                            full_service_data['region'])
                        recording_service_data['action'].append(
                            full_service_data['action'])
                        recording_service_data['full_name'].append(
                            f"{full_service_data['check_person']} | Телефон: {full_service_data['phone']}")
                        if full_service_data['status']:
                            recording_service_data['status'].append(
                                'Выполнено')
                        else:
                            recording_service_data['status'].append(
                                'Не выполнено')
                        recording_service_data['weather_condition'].append(
                            f"{full_service_data['par_name']}, от {full_service_data['action_date_begin']} до {full_service_data['action_date_end']}")
                    # change the value of 'UPLOADED' to 'SENT' to forbid furhter receipt
                    db.query(
                        f"select files_uploading.update_file_metadata('{file}')", f'')
                    read_and_create_file(
                        file_name, recording_service_data, file, upload_time.strftime(FORMAT_DATE))
except Exception as e:
    print('Exception', e)
else:
    print('Success!', datetime.today().strftime(FORMAT_DATE))
    curl.sent_message_to_express_chat()
